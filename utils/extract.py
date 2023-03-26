import csv
import openpyxl   


def sales_csv_to_list(path):
    list_sales_a = []
    with open(file=path, mode='r') as csv_file:
        csv_reader = csv.DictReader(csv_file, delimiter=",")

        for rec in csv_reader:
            list_sales_a.append(rec)

    return list_sales_a


def sales_xlsx_to_list(path):
    list_sales_b = []

    workbook = openpyxl.load_workbook(path)
    worksheet = workbook['Sheet1']

    header = ['id_transaksi',
              'tgl_transaksi',
              'id_pembeli',
              'id_produk',
              'potongan_harga']

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        sales_data = dict(zip(header, row))
        list_sales_b.append(sales_data)

    return list_sales_b


def brand_xlsx_to_list(path):
    list_brand = []

    workbook = openpyxl.load_workbook(path)
    worksheet = workbook['Sheet1']

    header = ["id_produk", "nama_brand", "name_produk", "harga"]
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        brand_dt = dict(zip(header, row))
        list_brand.append(brand_dt)

    return list_brand

